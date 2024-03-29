#!/usr/bin/python3
##
## Revise excel files
## http://github.com/Thaer-Maddah
##
## Copyright (C) 2023 Thaer Maddah. All rights reserved.

## Permission is hereby granted, free of charge, to any person obtaining
## a copy of this software and associated documentation files (the
## "Software"), to deal in the Software without restriction, including
## without limitation the rights to use, copy, modify, merge, publish,
## distribute, sublicense, and/or sell copies of the Software, and to
## permit persons to whom the Software is furnished to do so, subject to
## the following conditions:
##
## The above copyright notice and this permission notice shall be
## included in all copies or substantial portions of the Software.
##
## THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
## EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
## MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
## IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY
## CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
## TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
## SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
##
## [ MIT license: http://www.opensource.org/licenses/mit-license.php ]
##

import sys
import pandas as pd
import browse_files as bf
import openpyxl
import time
import re
import write_grades as wr


#excel_file='files/excel.xlsx'

grade = []


def check_file_name(file_name):
    if file_name == "وظيفة مهارات الحاسوب.xlsx" or file_name == "وظيفة مهارات حاسوب.xlsx" or \
                file_name == "مهارات الحاسوب.xlsx" or file_name == "مهارات حاسوب.xlsx":
        print(f"File name correct {file_name}")
        return True

    return False


def sheet_name(wb):
    title = '_'

    for ws in wb.worksheets:
        print(ws)
        match = re.search(title, ws.title, re.IGNORECASE)
        if match:
            print(f"Sheet name: {ws.title}")
            return True
        else:
            print(f"Sheet name is wrong: {ws.title}")
            return False

# This method find specific formula by substring
def eval_cell(ws, df, str_formula, str_value ):

    # Loop through each cell
    for row_index, row in df.iterrows():
        for col_index, cell_value in enumerate(row):
            cell = ws.cell(row=row_index + 1, column=col_index + 1)
            # Check if the cell has a formula
            if ( str(cell.value).find(str_formula) != -1):
                if type(str_value) is str:
                    match = re.search(str_value.strip(), cell_value, re.IGNORECASE)
                    if match is not None: # round the number for 2 decimals
                        print(f"Cell {str_formula} {cell.coordinate} contains a formula: {cell_value}")
                        return True
                else:
                    if round(cell_value, 0) == str_value: # round the number for 2 decimals
                        print(f"Cell {str_formula} {cell.coordinate} contains a formula: {cell_value}")
                        return True

    return False

def has_decimal(wbr):
    ws = wbr.active
    print(ws)
    # Get cell value and convert to float
    cell = ws['Q5'] # Replace 'A1' with the desired cell address
    print(cell.value)
    value = float(cell.value)
    
    # Check if value contains a decimal
    if int(value) == value:
        print(f"Cell {cell.coordinate} contains no decimal")
        return False
    else:
        print(f"Cell {cell.coordinate} contains a decimal")
        return True

# Charts
def check_excel_contains_chart(wb):
    
    has_chart = False
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if sheet._charts:
            has_chart = True
            break
    
    if has_chart:
        print("The Excel file contains a chart.")
    else:
        print("The Excel file does not contain a chart.")

    return has_chart


def get_chart_type(wb):
    
    chart_type = ""
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        if sheet._charts:
            charts = sheet._charts  # for only first chart type sheet._charts[0]
            
            # loop through sheet to check if there is more than one chart
            for chart in charts: 
                # get chart name
                chart_type = chart.__class__.__name__
                print("The chart type is:", chart_type)
            return chart_type
 

def has_formatting(ws):
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell has any conditional formatting
            for cf in ws.conditional_formatting:
                if cell.coordinate in cf:
                    # Print the cell coordinate and the conditional formatting rule
                    print(f"Cell {cell.coordinate} has conditional formatting: {cf.cfRule[0].dxfId}")
                    return True
    return False



def is_worksheet_empty(ws):
    """Check if a given worksheet is empty."""
    for row in ws.iter_rows(values_only=True):
        if any(cell for cell in row):
            return False
    return True




def reviseExcel(file_name, ws, wb, wbr, df):

    if check_file_name(file_name):
        grade.append(3)
    else:
        grade.append(0)

    if sheet_name(wb):
        grade.append(3)
    else:
        grade.append(0)


    str_dict = {
        "=CONCATENATE": 'يامن',
        "=DAY": 13,
        "=MONTH": 5,
        "=YEAR": 2000,
        "=SUM": 196,
        "=AVERAGE": 49,
        "=MAX": 60,
        "=MIN": 37,
        "=LARGE": 53,
        "=SMALL": 46,
        "=IF(": "D",
        "=COUNT": 15,
        "=COUNTIF": 7,
        "=COUNTIF(": 4
        }

    for key, value in str_dict.items():

        if eval_cell(ws, df, key, value ):
            if value == "D":
                grade.append(6)
            else:
                grade.append(2)
        else:
            grade.append(0)


    if has_decimal(wbr):
        grade.append(2)
    else:
        grade.append(0)


    if has_formatting(ws):
        grade.append(10)
    else:
        grade.append(0)


    if check_excel_contains_chart(wb):
        chart_type = get_chart_type(wb)
        if chart_type == "BarChart" or chart_type == "BarChart3D":
            grade.append(10)
        else:
            grade.append(5)
    else:
        grade.append(0)


    degree = sum(grade[1:len(grade)])
    
    print(grade)
    wr.writeExcelGrades([grade]) 
    print('Final degree is:', degree)
    del grade[:]
 

def main():
    folder = 'Assign/C14'
    ext = '.xlsx'
    trim_txt = '/mnt/c/code/Assign/'
    counter = 0 
    files = []
    files, dirs = bf.browse(ext, folder)
    start = time.time()
    for file, dir in zip(files, dirs):
        path = bf.getFile(file, dir)
        print(file)
        print(path)
        # openpyxl
        # data_only=True return the Cell value
        df = pd.read_excel(path, engine='openpyxl', header=None)
        wb = openpyxl.load_workbook(path, data_only=False)  # get cell formulas
        wbr = openpyxl.load_workbook(path, data_only=True)  # get cell value

        ws = wb.active
        print(f"Active worksheet: {ws}")
        grade.append(path.strip(trim_txt))
        reviseExcel(file, ws, wb, wbr, df)
    
        counter += 1
        print(counter, 'Excel file revised!')
    
        sep = '='
        print(sep*120)
        #time.sleep(1)
    
    end = time.time()
    print(f"Total time: {round(end - start)} seconds")


if __name__ == '__main__':
    sys.exit(main())
