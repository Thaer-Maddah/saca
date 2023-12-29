import openpyxl

def check_excel_contains_chart(filename):
    workbook = openpyxl.load_workbook(filename)
    
    has_chart = False
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet._charts:
            has_chart = True
            break
    
    if has_chart:
        print("The Excel file contains a chart.")
    else:
        print("The Excel file does not contain a chart.")

    return has_chart


def get_chart_type(filename):
    workbook = openpyxl.load_workbook(filename)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        if sheet._charts:
            charts = sheet._charts  # for only first chart type sheet._charts[0]
            
            # loop through sheet to check if there is more than one chart
            for chart in charts: 
                # get chart name
                chart_type = chart.__class__.__name__
                print("The chart type is:", chart_type)
            
            #return chart_type
    
    return None  # Return None if no chart is found


if __name__ == '__main__':
    if check_excel_contains_chart('test.xlsx'):
        # If there is chart in the file then get the type of chart
        get_chart_type('test.xlsx')
